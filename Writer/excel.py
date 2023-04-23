from openpyxl import Workbook
from Processor.excel import generate_excel_list, generate_despachos_list

def write_data(data, despachos):
    print('Writing...')
    wb = Workbook()
    ws = wb.create_sheet("ABRIL")
    wd = wb.create_sheet("DESPACHOS")

    list = generate_excel_list(data)
    list_despachos = generate_despachos_list(despachos)

    for row in list:
        ws.append(row)
    
    for row in list_despachos:
        wd.append(row)

    wb.save('./Output/test.xlsx')
    
    print('Complete :)')